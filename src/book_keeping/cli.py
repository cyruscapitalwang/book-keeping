from __future__ import annotations
import argparse
import re
from pathlib import Path
from datetime import datetime, date
from typing import List, Dict, Tuple, Optional
from math import isclose

from openpyxl import load_workbook
from openpyxl.styles import numbers, Alignment
from openpyxl.utils import get_column_letter


# ==============================
# PDF text extraction
# ==============================
def read_pdf_text(pdf_path: Path) -> str:
    """Extract text from PDF using pdfplumber if available, else PyPDF2."""
    try:
        import pdfplumber
        text = ""
        with pdfplumber.open(str(pdf_path)) as pdf:
            for p in pdf.pages:
                text += (p.extract_text(x_tolerance=1, y_tolerance=1) or "") + "\n"
        if text.strip():
            return text
    except Exception:
        pass
    try:
        from PyPDF2 import PdfReader
        rd = PdfReader(str(pdf_path))
        return "\n".join([(pg.extract_text() or "") for pg in rd.pages])
    except Exception:
        return ""


# ==============================
# Helpers / cleaners
# ==============================
# Allow amounts like ".21", "0.21", "47.15", "(1,204.21)", "-.59"
_AMOUNT_RE = r"[\(\-]?\$?(?:\d{1,3}(?:,\d{3})*|\d)?(?:\.\d{2})[\)]?"

def _norm_amount(s: str) -> Optional[float]:
    s1 = s.replace("$", "").replace(",", "").strip()
    neg = False
    if s1.startswith("(") and s1.endswith(")"):
        neg = True
        s1 = s1[1:-1]
    if s1.startswith("-"):
        neg = True
        s1 = s1[1:]
    # strip any trailing non-numeric ornament (e.g., diamond glyphs)
    s1 = re.sub(r"[^\d\.]+$", "", s1)
    try:
        v = float(s1)
    except ValueError:
        return None
    return -v if neg else v

def strip_after_first_amount(line: str) -> str:
    """Keep text only up to (and including) the FIRST amount on the line."""
    m = re.search(_AMOUNT_RE, line)
    if not m:
        return line
    return line[: m.end()].rstrip()

def clean_text(s: str) -> str:
    s = re.sub(r"\bPage\s+\d+\s+of\s+\d+\b", "", s, flags=re.I)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


# ==============================
# Checking parser (robust)
# ==============================
def parse_chase_transactions_full(
    text: str,
    default_year: int,
    verbose: bool = False,
) -> Tuple[List[Dict], Dict[str, float], Dict[str, List[str]]]:
    sect_deposits     = re.compile(r"^\s*DEPOSITS\s+AND\s+ADDITIONS\b", re.I)
    sect_withdrawals  = re.compile(r"^\s*ELECTRONIC\s+WITHDRAWALS\b", re.I)
    sect_fees         = re.compile(r"^\s*FEES\b", re.I)
    sect_daily_bal    = re.compile(r"DAILY\s+ENDING\s+BALANCE", re.I)
    line_total        = re.compile(r"^\s*TOTAL\b", re.I)

    date_re    = r"(0[1-9]|1[0-2])/(0[1-9]|[12]\d|3[01])(?:/(20\d{2}))?"
    amount_re  = _AMOUNT_RE

    tx_header_strict = re.compile(
        rf"^\s*{date_re}(?:\s+{date_re})?\s+(.*?)\s+({amount_re})\s*$"
    )
    tx_header_no_amt = re.compile(
        rf"^\s*{date_re}(?:\s+{date_re})?\s+(?!.*{amount_re}\s*$)(.+)$"
    )
    tx_header_loose  = re.compile(rf".*?{date_re}.*?({amount_re})\s*$")
    trailing_amt     = re.compile(rf"({amount_re})\s*$")

    transactions: List[Dict] = []
    pdf_totals = {"deposit": 0.0, "withdrawal": 0.0, "fee": 0.0}
    unparsed: Dict[str, List[str]] = {"deposit": [], "withdrawal": [], "fee": []}

    section = None
    current: Optional[Dict] = None
    pending_header: Optional[Dict] = None

    def finalize_current():
        nonlocal current
        if current is not None:
            current["Description"] = "\n".join(current["desc_lines"]).strip()
            transactions.append(current)
            current = None

    for raw in text.splitlines():
        ln = raw.strip()
        if sect_daily_bal.search(ln):
            finalize_current(); section = None; pending_header = None; continue
        
        if sect_deposits.search(ln):
            finalize_current(); section = "deposit"; pending_header = None; continue
        if sect_withdrawals.search(ln):
            finalize_current(); section = "withdrawal"; pending_header = None; continue
        if sect_fees.search(ln):
            finalize_current(); section = "fee"; pending_header = None; continue
        if not section:
            continue

        if line_total.search(ln):
            m_amt = trailing_amt.search(ln)
            if m_amt:
                val = _norm_amount(m_amt.group(1))
                if val is not None:
                    pdf_totals[section] = abs(val)
            finalize_current(); pending_header = None; continue

        # Two-line header: amount on next line
        if pending_header is not None:
            ln_clipped = strip_after_first_amount(ln)
            m_amt_first = re.search(_AMOUNT_RE, ln_clipped)
            if m_amt_first:
                amt = _norm_amount(m_amt_first.group(0))
                if amt is not None:
                    current = {
                        "Date": pending_header["date"],
                        "Amount": amt,
                        "Section": pending_header["section"],
                        "desc_lines": [pending_header["desc"], ln_clipped],
                    }
                    pending_header = None
                    continue
            pending_header["desc"] += " " + ln_clipped
            continue

        m = tx_header_strict.match(ln)
        if m:
            finalize_current()
            g = m.groups()
            d1m, d1d, d1y = g[0], g[1], g[2]
            d2m, d2d, d2y = g[3], g[4], g[5]
            desc = g[6]
            amt_str = g[7]
            year  = int((d2y or d1y) or str(default_year))
            month = int((d2m or d1m)); day = int((d2d or d1d))
            try:
                tx_date = datetime(year, month, day).date()
            except ValueError:
                unparsed[section].append(ln); continue
            amt = _norm_amount(amt_str)
            if amt is None:
                unparsed[section].append(ln); continue
            current = {"Date": tx_date, "Amount": amt, "Section": section, "desc_lines": [desc]}
            continue

        m_no = tx_header_no_amt.match(ln)
        if m_no:
            d1m, d1d, d1y = m_no.group(1), m_no.group(2), m_no.group(3)
            d2m, d2d, d2y = m_no.group(4), m_no.group(5), m_no.group(6)
            desc_only = m_no.group(7)
            year  = int((d2y or d1y) or str(default_year))
            month = int((d2m or d1m)); day = int((d2d or d1d))
            try:
                tx_date = datetime(year, month, day).date()
            except ValueError:
                unparsed[section].append(ln); continue
            finalize_current()
            pending_header = {"date": tx_date, "desc": desc_only, "section": section}
            continue

        m2 = tx_header_loose.match(ln)
        if m2:
            ln_clipped = strip_after_first_amount(ln)
            m_amt_first = re.search(_AMOUNT_RE, ln_clipped)
            dmatch = re.search(r"(0[1-9]|1[0-2])/(0[1-9]|[12]\d|3[01])(?:/(20\d{2}))?", ln_clipped)
            if not (m_amt_first and dmatch):
                unparsed[section].append(ln); continue
            amt = _norm_amount(m_amt_first.group(0))
            if amt is None:
                unparsed[section].append(ln); continue
            mm, dd, yy = int(dmatch.group(1)), int(dmatch.group(2)), dmatch.group(3)
            year = int(yy) if yy else default_year
            try:
                tx_date = datetime(year, mm, dd).date()
            except ValueError:
                unparsed[section].append(ln); continue
            finalize_current()
            current = {"Date": tx_date, "Amount": amt, "Section": section, "desc_lines": [ln_clipped]}
            continue

        if current is not None:
            current["desc_lines"].append(strip_after_first_amount(ln))
            continue

        unparsed[section].append(ln)

    finalize_current()
    if pending_header is not None:
        unparsed[pending_header["section"]].append(pending_header["desc"])
    return transactions, pdf_totals, unparsed


# ==============================
# Credit card parser (multi-line + balances)
# ==============================
_SECTION_STARTS = [
    r"ACCOUNT\s+ACTIVITY",
    r"TRANSACTIONS",
    r"TRANSACTION\s+ACTIVITY",
    r"CURRENT\s+CHARGES",
    r"ACTIVITY\s+DETAILS",
]

_SUMMARY_WITHIN_ACTIVITY = re.compile(
    r"^\s*(TRANSACTIONS\s+THIS\s+CYCLE|INCLUDING\s+PAYMENTS\s+RECEIVED|TOTAL|SUBTOTAL)\b",
    re.I
)

def _locate_cc_start(lines: List[str]) -> int:
    """Find index to start parsing from. Try known headers; else fall back to first date-at-start line."""
    for i, ln in enumerate(lines):
        for pat in _SECTION_STARTS:
            if re.search(rf"^\s*{pat}\b", ln, flags=re.I):
                return i + 1
    date_start = re.compile(r"^\s*(0[1-9]|1[0-2])/(0[1-9]|[12]\d|3[01])(?:/\d{2,4})?\b")
    for i, ln in enumerate(lines):
        if date_start.match(ln or ""):
            return i
    return 0

def parse_credit_card_transactions_multiline(
    text: str, default_year: int
) -> Tuple[List[Dict], Optional[float], Optional[float]]:
    """
    Header-agnostic CC parser:
      - Finds Previous/New balance anywhere before the first transaction line.
      - Treats ANY line containing a date token as a transaction start, even if the
        date isn't at the start of the line and even if it has a trailing footnote
        (e.g., 07/08/24* or 07/08/2024†).
      - Builds a block until the next date line or a summary row.
      - Extracts the amount from the block (last line first, then others).
    """
    lines = text.splitlines()

    # ---- balances (scan from the top until we see the first transaction line) ----
    date_token = re.compile(
        r"(?:^|\s)(?P<mm>0[1-9]|1[0-2])/"
        r"(?P<dd>0[1-9]|[12]\d|3[01])"
        r"(?:/(?P<yy>\d{2,4}))?"
        r"(?:[*•†‡])?(?=\D|$)",
        re.I,
    )
    amt_re = re.compile(_AMOUNT_RE)
    nb_re  = re.compile(r"\bNEW\s+BALANCE\b", re.I)
    pb_res = [
        re.compile(r"\bPREVIOUS\s+BALANCE\b", re.I),
        re.compile(r"\bBALANCE\s+FROM\s+LAST\s+STATEMENT\b", re.I),
        re.compile(r"\bPRIOR\s+BALANCE\b", re.I),
    ]
    summary_guard = re.compile(
        r"^\s*(TRANSACTIONS\s+THIS\s+CYCLE|INCLUDING\s+PAYMENTS\s+RECEIVED|TOTAL|SUBTOTAL)\b",
        re.I
    )

    new_balance: float | None = None
    prev_balance: float | None = None

    # find the first index that looks like a transaction line (has a date token)
    first_tx_idx = None
    for i, ln in enumerate(lines):
        if date_token.search(ln or ""):
            first_tx_idx = i
            break

    # grab balances from before that point
    scan_upto = first_tx_idx if first_tx_idx is not None else len(lines)
    for i in range(scan_upto):
        ln = lines[i]
        if new_balance is None and nb_re.search(ln):
            m_amt = amt_re.search(ln)
            if m_amt:
                v = _norm_amount(m_amt.group(0))
                if v is not None:
                    new_balance = v
        if prev_balance is None:
            for rx in pb_res:
                if rx.search(ln):
                    m_amt = amt_re.search(ln)
                    if m_amt:
                        v = _norm_amount(m_amt.group(0))
                        if v is not None:
                            prev_balance = v
                            break

    # ---- collect transactions (anywhere after the first tx-looking line) ----
    txs: List[Dict] = []
    if first_tx_idx is None:
        return txs, new_balance, prev_balance

    current_date: Optional[date] = None
    current_lines: List[str] = []

    def flush():
        nonlocal current_date, current_lines
        if current_date is None or not current_lines:
            current_date = None
            current_lines = []
            return
        # full description for rule-based categorization
        desc = clean_text(" ".join(current_lines))
        # extract an amount from the block (prefer the last line, then first, then others)
        amount_val: Optional[float] = None
        for line in (current_lines[-1], current_lines[0], *reversed(current_lines)):
            m = amt_re.search(line or "")
            if m:
                v = _norm_amount(m.group(0))
                if v is not None:
                    amount_val = v
                    break
        if amount_val is not None:
            txs.append({"Date": current_date, "Amount": amount_val, "Desc": desc})
        current_date = None
        current_lines = []

    for raw in lines[first_tx_idx:]:
        ln = (raw or "").rstrip()

        # summary/subtotal rows end a block but are themselves ignored
        if summary_guard.search(ln):
            flush()
            continue

        m = date_token.search(ln)
        if m:
            flush()
            mm = int(m.group("mm"))
            dd = int(m.group("dd"))
            yy = m.group("yy")
            if yy:
                yy = int(yy)
                yy = (2000 + yy) if yy < 100 else yy
            else:
                yy = default_year
            try:
                current_date = datetime(yy, mm, dd).date()
            except ValueError:
                current_date = None
            current_lines = [ln]
        else:
            if current_date is not None:
                current_lines.append(ln)

    flush()
    return txs, new_balance, prev_balance



# ==============================
# Excel helpers
# ==============================
def find_headers(ws):
    """Locate header row and key columns for the bank register sheet."""
    for r in range(1, 100):
        vals = [
            (str(ws.cell(row=r, column=c).value).strip().lower()
             if ws.cell(row=r, column=c).value is not None else "")
            for c in range(1, 41)
        ]
        date_col = next((i + 1 for i, v in enumerate(vals) if v in ("date", "transaction date", "posting date")), None)
        check_col = next((i + 1 for i, v in enumerate(vals) if v in ("check #", "check#", "check no", "check number")), None)
        amt_col = next((i + 1 for i, v in enumerate(vals) if v in ("amount", "amt", "amount (usd)", "debit/credit")), None)
        if date_col and check_col and amt_col:
            return r, date_col, check_col, amt_col
    return 1, 1, 2, 4


def compute_section_sums(transactions: List[Dict]) -> Dict[str, float]:
    sums = {"deposit": 0.0, "withdrawal": 0.0, "fee": 0.0}
    for t in transactions:
        sec = t["Section"]; amt = abs(float(t["Amount"]))
        sums[sec] += amt
    for k in list(sums.keys()):
        sums[k] = round(sums[k], 2)
    return sums


# ---------- Credit Card sheet helpers ----------
def scan_cc_blocks(ws) -> List[Dict]:
    """
    Find card blocks on 'Credit Card Register-Corp' and return:
      [{ 'top_cell': (r,c), 'digits': '0652', 'header_row': int, 'date_col': int, 'amount_col': int, 'cat_col': int }]
    More tolerant of label text (e.g., "Type of Cred") and uses a wider header search window.
    """
    label_ok = re.compile(r"type\s+of\s+cred", re.I)  # matches "Type of Credit Card", "Type of Cred", etc.
    blocks: List[Dict] = []

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            if not label_ok.search(v.strip()):
                continue

            # digits prefer right cell, fallback to same cell
            right_val = ws.cell(row=r, column=c + 1).value
            digits: Optional[str] = None
            if right_val is not None:
                m = re.search(r"(\d{4,5})\b", str(right_val))
                if m:
                    digits = m.group(1)
            if digits is None:
                m = re.search(r"(\d{4,5})\b", v)
                if m:
                    digits = m.group(1)
            if not digits:
                continue

            # Wider window for headers to the right of the label
            header_row = None
            date_col = amount_col = cat_col = None
            for rr in range(r + 1, min(r + 40, ws.max_row + 1)):  # was +25
                row_vals = [
                    (str(ws.cell(row=rr, column=cc).value).strip().lower()
                     if ws.cell(row=rr, column=cc).value is not None else "")
                    for cc in range(c, min(c + 40, ws.max_column + 1))  # was +16
                ]
                try:
                    d_idx = row_vals.index("date")
                    a_idx = row_vals.index("amount")
                    cat_idx = next(i for i, val in enumerate(row_vals)
                                   if val.startswith("expense category"))
                except (ValueError, StopIteration):
                    continue
                header_row = rr
                date_col = c + d_idx
                amount_col = c + a_idx
                cat_col = c + cat_idx
                break

            if header_row and date_col and amount_col and cat_col:
                blocks.append({
                    "top_cell": (r, c),
                    "digits": digits,
                    "header_row": header_row,
                    "date_col": date_col,
                    "amount_col": amount_col,
                    "cat_col": cat_col,
                })
            # Optional: debug if label found but headers not hooked
            # else:
            #     print(f"[DEBUG] Found label for {digits} at ({r},{c}) but could not locate headers nearby.")

    return blocks



def write_cc_block(ws, block: Dict, txs: List[Dict], previous_balance: Optional[float] = None):
    """
    Write txs into the block’s Date/Amount/Category columns starting at header_row+1.
    Do NOT delete/insert rows (avoids nuking other blocks on the same rows).
    Adds Expense Category classification per user rules.
    """
    start_row = block["header_row"] + 1
    date_col, amount_col, cat_col = block["date_col"], block["amount_col"], block["cat_col"]

    # 1) Clear a safe range in JUST THESE COLUMNS (e.g., next 600 rows)
    clear_rows = max(600, len(txs) + 50)
    for rr in range(start_row, start_row + clear_rows):
        ws.cell(row=rr, column=date_col).value = None
        ws.cell(row=rr, column=amount_col).value = None
        ws.cell(row=rr, column=cat_col).value = None

    def classify(desc: str, amt: float) -> str:
        d = (desc or "").lower()
        # (1) Negative amounts
        if amt < 0:
            if ("payment" in d) and (previous_balance is not None) and isclose(abs(amt), abs(float(previous_balance)), abs_tol=0.01):
                return "Automatic Payment Received From Checking Account"
            return "Office Supplies Credit"
        # (2) Office suppliers big-box
        if any(k in d for k in ["amazon", "walmart", "lowes", "costco", "home depot", "menards"]):
            return "Office Suppliers"
        # (3) Apple
        if "apple" in d:
            return "Company hardware purchase"
        # (4) Travel (transport/parking/etc.)
        if any(k in d for k in ["united", "iparkit", "driven car"]):
            return "Travel and Entertainment - Travel"
        # (5) Dues / subscriptions
        if any(k in d for k in ["chatgpt", "github", "yahoo"]):
            return "Dues and Subscriptions"
        # (6) Default
        return "Travel and Entertainment - Meals"

    # 2) Write transactions sequentially
    for i, t in enumerate(txs):
        r = start_row + i
        amt = float(t["Amount"])
        desc = t.get("Desc", "")
        ws.cell(row=r, column=date_col).value = t["Date"]
        ws.cell(row=r, column=amount_col).value = amt
        ws.cell(row=r, column=cat_col).value = classify(desc, amt)

    # 3) Formats for these columns only
    for i in range(max(1, len(txs))):
        r = start_row + i
        dcell = ws.cell(row=r, column=date_col)
        acell = ws.cell(row=r, column=amount_col)
        dcell.number_format = "MM/dd/yyyy"
        acell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # 4) Simple autofit for these columns only
    for col in (date_col, amount_col, cat_col):
        max_len = 0
        letter = get_column_letter(col)
        for r in range(block["header_row"], start_row + max(len(txs), 1)):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)


def find_pdf_for_digits(folder: Path, digits: str) -> Optional[Path]:
    """Prefer '*statements-<digits>*.pdf', else any '*<digits>*.pdf'."""
    preferred = sorted(folder.glob(f"*statements-{digits}*.pdf"))
    if preferred:
        return preferred[0]
    anymatch = sorted(folder.glob(f"*{digits}*.pdf"))
    return anymatch[0] if anymatch else None


# ==============================
# Main
# ==============================
def main():
    parser = argparse.ArgumentParser(
        description="Book-keeping: build/update registers. Template is read from project root."
    )
    parser.add_argument("--directory", required=True,
                        help="Target folder named YYYY-MM with bank (*2590*.pdf) and card PDFs (filenames contain last digits).")
    parser.add_argument("--dry-run", action="store_true", help="Simulate; do not write Excel")
    parser.add_argument("--verbose", action="store_true", help="Print detailed steps")
    parser.add_argument("--force", action="store_true", help="Write Excel even if totals mismatch")
    args = parser.parse_args()

    target = Path(args.directory)
    if not target.exists() or not target.is_dir():
        raise FileNotFoundError(f"Directory not found: {target}")

    # folder YYYY-MM → year/month
    month_tag = target.name
    m = re.fullmatch(r"(20\d{2})-(0[1-9]|1[0-2])", month_tag)
    if not m:
        raise ValueError(f"Folder name must be YYYY-MM (got: {month_tag})")
    year = int(m.group(1)); month = int(m.group(2))
    yyyymm = f"{year}{month:02d}"
    first_day = datetime(year, month, 1)

    # template at project root
    project_root = Path(__file__).resolve().parent.parent.parent
    template = project_root / "Corp Registers_.xlsx"
    if not template.exists():
        raise FileNotFoundError(
            f"Excel template not found at project root: {template}\n"
            "Place 'Corp Registers_.xlsx' next to your pyproject.toml."
        )

    # pick checking PDF (*2590*)
    checking_pdf = None
    for cand in sorted(target.glob("*.pdf")):
        if "2590" in cand.name:
            checking_pdf = cand; break

    if args.verbose:
        print(f"[INFO] Project root: {project_root}")
        print(f"[INFO] Month tag: {month_tag} -> {yyyymm}")
        print(f"[INFO] Template: {template.name}")
        if checking_pdf: print(f"[INFO] Checking PDF: {checking_pdf.name}")

    dest_excel = target / f"Corp Registers_{yyyymm}.xlsx"
    if not args.dry_run:
        if template.resolve() != dest_excel.resolve():
            dest_excel.write_bytes(template.read_bytes())

    wb = load_workbook(dest_excel)

    # --------- Bank register population ----------
    if checking_pdf:
        text = read_pdf_text(checking_pdf)
        transactions, pdf_totals, unparsed = parse_chase_transactions_full(text, default_year=year, verbose=args.verbose)

        ordered_tx: List[Dict] = []
        for sec in ("deposit", "withdrawal", "fee"):
            ordered_tx.extend([t for t in transactions if t["Section"] == sec])

        sums_from_data = compute_section_sums(ordered_tx)

        print("PDF Section Totals:")
        print(f"  Deposits   : ${pdf_totals['deposit']:.2f}")
        print(f"  Withdrawals: ${pdf_totals['withdrawal']:.2f}")
        print(f"  Fees       : ${pdf_totals['fee']:.2f}")
        print("Computed Totals (from data):")
        print(f"  Deposits   : ${sums_from_data['deposit']:.2f}")
        print(f"  Withdrawals: ${sums_from_data['withdrawal']:.2f}")
        print(f"  Fees       : ${sums_from_data['fee']:.2f}")

        mismatches = [sec for sec in ("deposit", "withdrawal", "fee")
                      if not isclose(pdf_totals[sec], sums_from_data[sec], abs_tol=0.01)]
        if mismatches and not args.force:
            raise AssertionError("Section totals mismatch between PDF and parsed data. Use --force to proceed.")
        elif mismatches:
            print("[WARN] Totals mismatch, but --force specified; continuing.")

        ws = wb["Check Register-Corp"]
        ws["B9"].value = first_day
        header_row, date_col, check_col, amt_col = find_headers(ws)
        expense_col = 4; deposit_col = 5

        if header_row < ws.max_row:
            ws.delete_rows(header_row + 1, ws.max_row - header_row)

        for r_i, t in enumerate(ordered_tx, start=header_row + 1):
            ws.cell(row=r_i, column=date_col).value = t["Date"]
            ws.cell(row=r_i, column=amt_col).value = abs(float(t["Amount"]))
            full_desc = clean_text((t.get("Description") or "").replace("\r", " ").replace("\n", " ").replace("\t", " "))
            ws.cell(row=r_i, column=check_col).value = full_desc
            ws.cell(row=r_i, column=check_col).alignment = Alignment(wrap_text=True)

            dlow = full_desc.lower()
            amt_abs = round(abs(float(t["Amount"])), 2)

            if t["Section"] == "deposit":
                if "transfer" in dlow and "0639" in dlow:
                    ws.cell(row=r_i, column=deposit_col).value = "Promissory Note to bond holder Ting Wang"
                elif "e*trade" in dlow:
                    ws.cell(row=r_i, column=deposit_col).value = "Transfer money from E*Trade Brokerage Account"
                elif "gainsystems" in dlow:
                    ws.cell(row=r_i, column=deposit_col).value = "Income by Consulting with GAINSystems"
                elif "allegis group" in dlow:
                    ws.cell(row=r_i, column=deposit_col).value = "Income by Consulting with JP Morgan Chase"
                elif "quinnox" in dlow:
                    ws.cell(row=r_i, column=deposit_col).value = "Income by Consulting with US Bank"
                else:
                    ws.cell(row=r_i, column=deposit_col).value = "Income"
            else:
                value = None
                if ("santander" in dlow):
                    value = "Car lease payment"
                elif ("telsa finasec" in dlow) and ("ting wang" in dlow):
                    value = "Car lease payment"
                elif ("chase credit crd" in dlow) and ("autopaybuss" in dlow):
                    value = "Credit Card Payment"
                elif ("payment to chase card" in dlow):
                    value = "Credit card payment"
                elif ("monthly service fee" in dlow):
                    value = "Bank service fee"
                elif ("e*trade" in dlow) and (amt_abs in (4617.50, 1250.00)):
                    value = "Xuefen Xie 401K contribution"
                if value is None:
                    if "e*trade" in dlow:
                        value = "Transfer money to E*Trade Brokerage Account"
                    elif "transfer" in dlow and "0639" in dlow:
                        value = "Return money to bond holder Ting Wang"
                    elif "gusto" in dlow:
                        value = "Payroll professional service fee"
                    elif (("u.s. bank" in dlow) or ("us bank" in dlow)) and (("lse pmts" in dlow) or ("lease" in dlow)):
                        value = "Car lease payment"
                    elif ("tesla" in dlow) or ("telsa" in dlow):
                        value = "Car wireless subscription payment"
                    else:
                        value = "Payment"
                ws.cell(row=r_i, column=expense_col).value = value

        # format & hide extras
        for rr in range(header_row + 1, ws.max_row + 1):
            dcell = ws.cell(row=rr, column=date_col)
            if isinstance(dcell.value, (datetime, date)):
                dcell.number_format = "MM/dd/yyyy"
            acell = ws.cell(row=rr, column=amt_col)
            if isinstance(acell.value, (int, float)):
                acell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

        for name in wb.sheetnames:
            lname = name.strip().lower()
            if lname in ("information to provide to ta", "typical exp in trading business"):
                wb[name].sheet_state = "hidden"

        for c in range(1, ws.max_column + 1):
            max_len = 0
            letter = get_column_letter(c)
            for rr in range(1, ws.max_row + 1):
                v = ws.cell(row=rr, column=c).value
                if v is None: continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[letter].width = min(max_len + 2, 80)

    # --------- Credit Card Register-Corp ----------
    if "Credit Card Register-Corp" in wb.sheetnames:
        ws_cc = wb["Credit Card Register-Corp"]
        blocks = scan_cc_blocks(ws_cc)
        print(f"[INFO] Credit-card blocks found: {len(blocks)}")
        for blk in blocks:
            digits = blk["digits"]
            print(f"[INFO] Block digits: {digits} at header row {blk['header_row']}")
            pdf = find_pdf_for_digits(target, digits)
            if not pdf:
                print(f"[WARN] No PDF found in folder for card *{digits}. Skipping.")
                continue
            print(f"[INFO] Card {digits}: using PDF {pdf.name}")

            text = read_pdf_text(pdf)
            cc_txs, new_balance, previous_balance = parse_credit_card_transactions_multiline(text, default_year=year)

            print(f"[INFO] Card {digits}: parsed {len(cc_txs)} transactions.")

            # Generic validation: Previous + Σ(all amounts) == New
            sum_all = round(sum(float(t["Amount"]) for t in cc_txs), 2)

            nb_str = f"${new_balance:.2f}" if new_balance is not None else "(not found)"
            pb_str = f"${previous_balance:.2f}" if previous_balance is not None else "(not found)"
            print(f"[CARD {digits}] Previous Balance (PDF): {pb_str}")
            print(f"[CARD {digits}] New Balance (PDF)    : {nb_str}")
            print(f"[CARD {digits}] Σ(all amounts)       : ${sum_all:.2f}")
            print(f"[CARD {digits}] Check: Previous + Σ(amounts) == New")

            if (new_balance is None) or (previous_balance is None):
                missing = []
                if previous_balance is None: missing.append("Previous")
                if new_balance is None:     missing.append("New")
                msg = f"Missing {' and '.join(missing)} balance(s) for *{digits}*; cannot validate."
                if not args.force:
                    raise AssertionError(msg)
                else:
                    print("[WARN]", msg, "— continuing due to --force.")
            else:
                lhs = round(float(previous_balance) + sum_all, 2)
                rhs = round(float(new_balance), 2)
                if not isclose(lhs, rhs, abs_tol=0.01):
                    msg = (f"Validation failed for *{digits}*: "
                           f"Previous (${previous_balance:.2f}) + Σ(${sum_all:.2f}) "
                           f"= ${lhs:.2f} != New (${new_balance:.2f})")
                    if not args.force:
                        raise AssertionError(msg)
                    else:
                        print("[WARN]", msg, "— continuing due to --force.")

            # Write the block (now with category rules; pass previous_balance)
            write_cc_block(ws_cc, blk, cc_txs, previous_balance=previous_balance)
    else:
        print("[WARN] Sheet 'Credit Card Register-Corp' not found; skipping CC population.")

    if not args.dry_run:
        wb.save(dest_excel)
        print(f"Workbook updated: {dest_excel}")
    else:
        print("[DRY-RUN] No files written.")


if __name__ == "__main__":
    main()
